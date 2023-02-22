import openpyxl

def findFCS(PythonStub):

    sheet_obj = sheet['Parser Q']
    row = sheet_obj.max_row #row length
    column = sheet_obj.max_column #col length
    print("Total Rows:", row)
    print("Total Columns:", column)

    right_fcs=[]
    wrong_fcs = []
    for j in range(2, row): #a loop that goes through all the rows
        isChanges = False #indicates changes rows
        sum = 0
        for i in range(2, 21): #a loop through all the col ,the range is the header and data bytes range
            cell_obj = sheet_obj.cell(row=j, column=i)
            print(cell_obj.value, end=" ") #cell value
            if(check7E(sheet,j,i)==True) : # function that checks if there "7D" and then "5E"
                isChanges=True
                sum+=int("7E",16)
                continue
            elif(check7E(sheet,j,i-1)==True): #skip 7D-5E col
                continue
            elif (check7D(sheet, j, i) == True): #function that checks if there "7D" and then "5D"
                isChanges = True
                sum += int("7D", 16)
                continue
            elif (check7D(sheet, j, i - 1) == True):#skip 7D-5D col
                continue
            else :
                sum += int(cell_obj.value, 16) #convert hex2dec
        temp = hex(255 - (sum % 256))[2:].upper() #fcs formula
        num=Convert(sum,temp) #function that converts values like 3 to 03 for comparsion
        print("FCS to hex is" + " " +num)
        if(isChanges==False): #condition for case that 7D-5E or 7D-5D at the end of data stream
            if(check7E(sheet,j,21)==True):
                if(num!="7E"): ##if 7D-5E instead of FCS val then compare to 7E
                    wrong_fcs.append('row=' + str(j) + " ," + sheet_obj.cell(row=j, column=21).value)
                    right_fcs.append('row=' + str(j) + " ," + temp)
            elif(check7D(sheet,j,21)==True) :
                if(num!="7D"):
                    wrong_fcs.append('row=' + str(j) + " ," + sheet_obj.cell(row=j, column=21).value)
                    right_fcs.append('row=' + str(j) + " ," + temp)
            elif (num != sheet_obj.cell(row=j, column=21).value):
                wrong_fcs.append('row=' + str(j) + " ," + sheet_obj.cell(row=j, column=21).value )
                right_fcs.append('row=' + str(j) + " ," + temp)
        else:
            if(num != sheet_obj.cell(row=j, column=22).value):
                wrong_fcs.append('row=' + str(j) + " ," + sheet_obj.cell(row=j, column=21).value)
                right_fcs.append('row=' + str(j) + " ," + temp)
        print('The FCS val is' + " " + str(255 - (sum % 256)))
    print("The wrong FCS is :", wrong_fcs)
    print("The right FCS is :", right_fcs)

def Convert (sum,temp) :
    if (255 - (sum % 256)< 16):
        return '0' + temp
    else :
        return temp
def check7E(PythonStub,j,i) :
    sheet_obj=sheet['Parser Q']
    if(sheet_obj.cell(row=j, column=i).value=="7D" and sheet_obj.cell(row=j, column=i+1).value=="5E"):
        return True
    else:
        return False
def check7D(PythonStub,j,i) :
    sheet_obj=sheet['Parser Q']
    if(sheet_obj.cell(row=j, column=i).value=="7D" and sheet_obj.cell(row=j, column=i+1).value=="5D"):
        return True
    else:
        return False

if __name__ == '__main__':
    path = "Copy of Data_analysis_Q - V2.4.xlsx"
    sheet=openpyxl.load_workbook(path)

    findFCS(sheet)

